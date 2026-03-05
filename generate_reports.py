"""
Auralith Digital — Daily Report Generator
Generates 30 audit report HTML files from MedSpaOutreach_v3.xlsx
AI content (revenue leaks + outreach hook) generated at build time via Anthropic API.

Usage:
  ANTHROPIC_API_KEY=your_key python generate_reports.py

Output: /reports/ folder ready to push to GitHub Pages
"""

import openpyxl
import os
import re
import csv
import json
import time
import urllib.request
from datetime import date

TEMPLATE_PATH = 'report_template.html'
OUTPUT_DIR = 'docs/medspa'
REPORTS_PER_DAY = 30
API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def slugify(name):
    name = name.lower()
    name = re.sub(r'[^a-z0-9\s-]', '', name)
    name = re.sub(r'\s+', '-', name.strip())
    return re.sub(r'-+', '-', name)[:60]

def format_followers(n):
    if not n or n == 0: return 'Not found'
    if n >= 1000: return f'{n/1000:.1f}K'
    return str(n)

def review_signal(c):
    if c > 500: return 'Strong social proof'
    if c > 150: return 'Good trust baseline'
    if c > 50:  return 'Building credibility'
    if c > 0:   return 'Low visibility risk'
    return 'No reviews found'

def rating_signal(r):
    if not r: return 'No rating'
    r = float(r)
    if r >= 4.8: return 'Excellent reputation'
    if r >= 4.5: return 'Strong reputation'
    if r >= 4.0: return 'Average reputation'
    return 'Reputation risk'

def followers_signal(n):
    if not n or n == 0: return 'No Instagram data'
    if n > 10000: return 'Strong social presence'
    if n > 2000:  return 'Growing audience'
    if n > 500:   return 'Limited reach'
    return 'Very low reach'

def score_class(s):
    if s >= 70: return 'good'
    if s >= 45: return 'warn'
    return 'bad'

def calc_scores(b):
    vis = 15
    if b['reviews'] > 150: vis += 20
    elif b['reviews'] >= 75: vis += 15
    else: vis += 5
    if b['followers'] > 5000: vis += 25
    elif b['followers'] >= 1000: vis += 15
    elif b['followers'] > 0: vis += 8
    conv = 30 if b['website'] else 10
    mon = 15
    return {'vis': min(vis,100), 'conv': min(conv,100), 'mon': mon,
            'overall': round((min(vis,100)+min(conv,100)+mon)/3)}


# ─── AI GENERATION ────────────────────────────────────────────────────────────

def generate_ai_content(b):
    if not API_KEY:
        return fallback_content(b)

    prompt = f"""You are a digital marketing analyst for Auralith Digital, an automation agency for med spas in South Florida.

BUSINESS:
- Name: {b['name']}
- City: {b['city']}, FL
- Instagram: {b['ig_handle'] or 'Not found'} ({b['followers']:,} followers)
- Google: {b['reviews']} reviews, {b['rating']} stars
- Website: {b['website'] or 'None found'}

Respond ONLY with valid JSON (no markdown, no explanation):
{{
  "leaks": [
    {{
      "severity": "high|medium|low",
      "title": "5 words max",
      "description": "One sentence. Specific problem and why it costs money.",
      "estimate": "$X–Y/mo"
    }}
  ],
  "hook": "2-3 sentences. Use their actual numbers. Direct and specific. No generic phrases."
}}

Generate 3-5 leaks. Base them on: low followers=visibility gap, no website=conversion gap, low reviews=trust gap, assume no SMS/email/membership."""

    payload = json.dumps({
        'model': 'claude-sonnet-4-20250514',
        'max_tokens': 900,
        'messages': [{'role': 'user', 'content': prompt}]
    }).encode('utf-8')

    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=payload,
        headers={
            'Content-Type': 'application/json',
            'x-api-key': API_KEY,
            'anthropic-version': '2023-06-01'
        }
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = json.loads(resp.read())
            text = re.sub(r'^```json|```$', '', data['content'][0]['text'].strip(), flags=re.MULTILINE).strip()
            return json.loads(text)
    except Exception as e:
        print(f" [fallback: {e}]", end='')
        return fallback_content(b)

def fallback_content(b):
    leaks = []
    if b['followers'] < 2000:
        leaks.append({'severity':'high','title':'Low Instagram Reach',
            'description':"With under 2,000 followers, new patients in your area can't discover you through social search.",
            'estimate':'$1,000–2,000/mo'})
    if b['reviews'] < 100:
        leaks.append({'severity':'high','title':'Thin Review Count',
            'description':'Fewer than 100 Google reviews puts you below the trust threshold most new patients expect.',
            'estimate':'$800–1,500/mo'})
    if not b['website']:
        leaks.append({'severity':'high','title':'No Website Found',
            'description':"Without a website, every patient who Googles you hits a dead end before booking.",
            'estimate':'$1,500–3,000/mo'})
    leaks.append({'severity':'medium','title':'No Retention System',
        'description':"Without SMS or email follow-up, lapsed clients have no path back to rebook.",
        'estimate':'$600–1,200/mo'})
    leaks.append({'severity':'low','title':'No Membership Program',
        'description':"Membership programs lock in predictable monthly revenue and reduce churn.",
        'estimate':'$400–800/mo'})
    hook = (f"{b['name']} has {b['reviews']} Google reviews and "
            f"{format_followers(b['followers'])} Instagram followers. "
            f"Without automated follow-up and a clear retention system, "
            f"you're leaving returning client revenue on the table every month.")
    return {'leaks': leaks[:4], 'hook': hook}


# ─── HTML RENDERING ───────────────────────────────────────────────────────────

def render_leak(leak):
    colors = {
        'high':   ('#dc2626','#fef2f2','#fecaca'),
        'medium': ('#d97706','#fffbeb','#fde68a'),
        'low':    ('#6b7280','#f9fafb','#e5e7eb'),
    }
    badge, bg, border = colors.get(leak['severity'], colors['low'])
    return f"""<div style="display:flex;align-items:flex-start;gap:16px;padding:20px 24px;border-radius:8px;background:{bg};border:1px solid {border};margin-bottom:12px;">
  <span style="font-size:9px;font-weight:700;letter-spacing:0.1em;text-transform:uppercase;padding:3px 8px;border-radius:4px;background:{badge};color:white;white-space:nowrap;margin-top:2px;flex-shrink:0;">{leak['severity'].upper()}</span>
  <div>
    <div style="font-weight:600;font-size:15px;margin-bottom:4px;color:#0f0f0f;">{leak['title']}</div>
    <div style="font-size:14px;color:#6b7280;line-height:1.6;">{leak['description']}</div>
    <div style="font-size:13px;font-weight:600;color:{badge};margin-top:6px;">{leak['estimate']} in estimated monthly loss</div>
  </div>
</div>"""

def render_score_card(label, value, cls):
    bar = {'good':'#16a34a','warn':'#d97706','bad':'#dc2626','overall':'#0f0f0f'}[cls]
    txt = {'good':'#16a34a','warn':'#d97706','bad':'#dc2626','overall':'#0f0f0f'}[cls]
    return f"""<div style="background:#fff;padding:28px 24px;text-align:center;">
  <div style="font-size:10px;font-weight:600;letter-spacing:0.1em;text-transform:uppercase;color:#6b7280;margin-bottom:10px;">{label}</div>
  <div style="font-family:'DM Serif Display',serif;font-size:44px;line-height:1;color:{txt};margin-bottom:6px;">{value}</div>
  <div style="height:4px;background:#e5e7eb;border-radius:2px;margin-top:8px;"><div style="height:100%;width:{value}%;background:{bar};border-radius:2px;"></div></div>
</div>"""

def generate_report(b, template, ai):
    today = date.today().strftime('%B %d, %Y')
    slug = slugify(b['name'])
    scores = calc_scores(b)

    leaks_html = ''.join(render_leak(l) for l in ai['leaks'])
    high_count = sum(1 for l in ai['leaks'] if l['severity'] == 'high')
    tag_color = '#dc2626' if high_count >= 2 else '#d97706'
    tag_bg = '#fef2f2' if high_count >= 2 else '#fffbeb'
    leak_tag = f'<span style="font-size:11px;font-weight:600;letter-spacing:0.08em;text-transform:uppercase;padding:4px 10px;border-radius:20px;background:{tag_bg};color:{tag_color};">{len(ai["leaks"])} gaps found</span>'

    score_cards = (render_score_card('Overall Score', scores['overall'], 'overall') +
                   render_score_card('Visibility', scores['vis'], score_class(scores['vis'])) +
                   render_score_card('Conversion', scores['conv'], score_class(scores['conv'])) +
                   render_score_card('Monetization', scores['mon'], score_class(scores['mon'])))

    booking_val = 'Needs Check' if b['website'] else 'Missing'
    booking_sig = 'Website found — path unverified' if b['website'] else 'No website found'

    ig_display = b['ig_handle'] if b['ig_handle'] else 'Not found'

    html = template
    replacements = {
        '{{BUSINESS_NAME}}': b['name'],
        '{{BUSINESS_NAME_URL}}': b['name'].replace(' ', '%20'),
        '{{CITY}}': b['city'],
        '{{WEBSITE}}': b['website'],
        '{{IG_HANDLE}}': ig_display,
        '{{IG_FOLLOWERS}}': format_followers(b['followers']),
        '{{IG_FOLLOWERS_DISPLAY}}': format_followers(b['followers']),
        '{{GOOGLE_REVIEWS}}': str(b['reviews']) if b['reviews'] else '0',
        '{{GOOGLE_RATING}}': str(b['rating']) if b['rating'] else 'N/A',
        '{{AUDIT_DATE}}': today,
        '{{REVIEW_SIGNAL}}': review_signal(b['reviews']),
        '{{RATING_SIGNAL}}': rating_signal(b['rating']),
        '{{FOLLOWERS_SIGNAL}}': followers_signal(b['followers']),
        '{{SCORE_CARDS}}': score_cards,
        '{{LEAK_COUNT_TAG}}': leak_tag,
        '{{LEAKS_HTML}}': leaks_html,
        '{{HOOK_TEXT}}': ai['hook'],
        '{{BOOKING_VALUE}}': booking_val,
        '{{BOOKING_SIGNAL}}': booking_sig,
    }
    for k, v in replacements.items():
        html = html.replace(k, str(v))
    return slug, html


# ─── LOAD BUSINESSES ──────────────────────────────────────────────────────────

def load_businesses(xlsx_path):
    wb = openpyxl.load_workbook(xlsx_path)
    lt = wb['Lead Tracker']
    businesses = []
    for r in range(4, 800):
        name = lt.cell(row=r, column=1).value
        if not name: break
        if lt.cell(row=r, column=31).value: continue  # already contacted
        try: reviews = int(lt.cell(row=r, column=12).value)
        except: reviews = 0
        try: rating = float(lt.cell(row=r, column=13).value)
        except: rating = 0.0
        try: followers = int(lt.cell(row=r, column=6).value)
        except: followers = 0
        ig = str(lt.cell(row=r, column=5).value or '').strip()
        website = str(lt.cell(row=r, column=3).value or '').strip()
        priority = (reviews*0.5)+(followers*0.3)+(rating*10)+(50 if ig else 0)+(30 if website else 0)
        businesses.append({
            'row': r, 'name': str(name).strip(),
            'city': str(lt.cell(row=r, column=2).value or '').strip(),
            'website': website, 'phone': str(lt.cell(row=r, column=4).value or '').strip(),
            'ig_handle': ig, 'followers': followers, 'reviews': reviews,
            'rating': rating, 'priority': priority,
        })
    businesses.sort(key=lambda x: x['priority'], reverse=True)
    return businesses


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    xlsx_path = 'MedSpaOutreach_v3.xlsx'
    if not os.path.exists(xlsx_path):
        xlsx_path = '/mnt/user-data/outputs/MedSpaOutreach_v3.xlsx'

    if not API_KEY:
        print("⚠  ANTHROPIC_API_KEY not set — using fallback content")
        print("   export ANTHROPIC_API_KEY=sk-ant-...\n")

    print(f"Loading from {xlsx_path}...")
    businesses = load_businesses(xlsx_path)
    print(f"Found {len(businesses)} uncontacted businesses\n")

    with open(TEMPLATE_PATH, 'r') as f:
        template = f.read()

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    batch = businesses[:REPORTS_PER_DAY]
    index_rows = []

    for i, b in enumerate(batch, 1):
        print(f"  [{i:02d}/{len(batch)}] {b['name'][:45]}", end='', flush=True)
        ai = generate_ai_content(b)
        slug, html = generate_report(b, template, ai)
        with open(os.path.join(OUTPUT_DIR, f'{slug}.html'), 'w', encoding='utf-8') as f:
            f.write(html)
        url = f'https://reports.auralithdigital.com/medspa/{slug}.html'
        index_rows.append({'name':b['name'],'city':b['city'],'ig':b['ig_handle'],'reviews':b['reviews'],'url':url})
        print(f" ✓")
        time.sleep(0.3)

    with open('todays_outreach.csv', 'w', newline='') as f:
        csv.DictWriter(f, fieldnames=['name','city','ig','reviews','url']).writeheader()
        csv.DictWriter(f, fieldnames=['name','city','ig','reviews','url']).writerows(index_rows)

    print(f"\n✅ {len(batch)} reports → /{OUTPUT_DIR}/")
    print(f"✅ Outreach list → todays_outreach.csv")
    print(f"\ngit add reports/ && git commit -m 'Batch {date.today()}' && git push")

if __name__ == '__main__':
    main()
