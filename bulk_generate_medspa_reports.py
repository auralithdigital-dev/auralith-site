"""
Bulk generator: 150 med spa audit report HTML files.
Reads MedSpaOutreach_v3.xlsx (Lead Tracker sheet, headers row 3, data row 4+).
Sorts all businesses by review count descending.
Skips businesses already listed in todays_outreach.csv or with existing subdirectory.
Generates the next 150 as docs/medspa/[slug]/index.html.
"""

import os
import re
import csv
import openpyxl
from datetime import date

XLSX_PATH   = 'MedSpaOutreach_v3.xlsx'
CSV_PATH    = 'todays_outreach.csv'
OUTPUT_BASE = 'docs/medspa'
AUDIT_DATE  = 'March 8, 2026'
TARGET      = 150


# ── Helpers ───────────────────────────────────────────────────────────────────

def slugify(name):
    name = name.lower()
    name = re.sub(r'[^a-z0-9\s-]', '', name)
    name = re.sub(r'\s+', '-', name.strip())
    return re.sub(r'-+', '-', name)[:60]

def clean_ig(handle):
    if not handle: return ''
    handle = str(handle).strip().lstrip('@')
    return f'@{handle}' if handle else ''

def safe_int(val):
    try:    return int(val) if val else 0
    except: return 0

def safe_float(val):
    try:    return float(val) if val else 0.0
    except: return 0.0


# ── Content generators ────────────────────────────────────────────────────────

def gen_s01(name, city, reviews, rating):
    """01 — What's Working"""
    if rating >= 5.0:
        return (
            f"A perfect {rating:.1f}-star rating across {reviews} reviews is a remarkable "
            f"achievement for an aesthetics practice in {city}. That level of client trust "
            f"makes every follow-up and retention effort significantly more effective because "
            f"the goodwill is already established and easy to build on."
        )
    elif rating >= 4.8:
        return (
            f"A {rating:.1f}-star rating across {reviews} reviews reflects consistent client "
            f"satisfaction and strong clinical credibility in {city}. This reputation is the "
            f"foundation that makes every marketing and retention effort more effective, and "
            f"it deserves a system that actively protects and expands it."
        )
    elif rating >= 4.5:
        return (
            f"{name} has built {reviews} Google reviews and a {rating:.1f}-star rating in "
            f"{city}, giving the practice a solid trust baseline that most local competitors "
            f"cannot match. That social proof is most powerful when paired with automated "
            f"systems that keep clients engaged and returning between visits."
        )
    else:
        return (
            f"{name} has accumulated {reviews} Google reviews in {city}, giving prospective "
            f"clients a meaningful body of evidence to evaluate before booking. The volume of "
            f"reviews itself is a trust signal, and there is real opportunity to improve the "
            f"rating while building systems that make consistent review collection automatic."
        )


def gen_s02(name, city, reviews):
    """02 — Missed Bookings Without Follow-Up"""
    if reviews > 300:
        return (
            f"Prospective clients who discover {name} through Instagram or Google and leave "
            f"without booking are not being captured into any automated follow-up sequence. "
            f"In the competitive {city} aesthetics market, warm leads who do not hear back "
            f"within 24 to 48 hours routinely book with whoever follows up first. A practice "
            f"with {reviews} reviews already has the trust signals to convert those leads at "
            f"a high rate, but only if a system is actively working to reach them before they "
            f"choose someone else."
        )
    else:
        return (
            f"Prospective clients who find {name} online and leave without booking are not "
            f"entering any automated follow-up sequence. In {city}'s aesthetics market, warm "
            f"leads who do not hear back within 48 hours typically book with competitors who "
            f"follow up first. Without a capture and follow-up system, a meaningful share of "
            f"interested prospects goes unconverted every week."
        )


def gen_s03(name, city, reviews, rating):
    """03 — Inactive Client Leakage"""
    if reviews > 400:
        return (
            f"With {reviews} reviews representing a large and loyal client base, there is a "
            f"substantial group of past clients who have not returned recently and are not "
            f"receiving any proactive outreach to bring them back. Aesthetics clients go "
            f"inactive due to life disruptions rather than dissatisfaction, and a warm "
            f"personalized reactivation SMS is consistently the fastest revenue-generating "
            f"action for a practice of this size in {city}."
        )
    elif reviews > 150:
        return (
            f"With {reviews} past clients represented in Google reviews, there is a meaningful "
            f"group of people who had a positive experience at {name} and have not returned. "
            f"Without automated reactivation outreach, those clients quietly drift to other "
            f"options over time. An automated sequence targeting clients inactive for 90 days "
            f"can recover a significant number of appointments each month at very low cost."
        )
    else:
        return (
            f"{name} has a base of past clients who visited and have not been systematically "
            f"re-engaged. Without a process to identify and reach out to those clients, the "
            f"practice relies entirely on people remembering to rebook on their own. Automated "
            f"reactivation outreach is one of the highest-return investments for any practice "
            f"at this stage of growth in {city}."
        )


def gen_s04(name, city, reviews, rating):
    """04 — No Automated Review Generation"""
    if rating >= 4.8:
        return (
            f"Maintaining a {rating:.1f}-star average requires a steady flow of new reviews "
            f"to protect the rating as client volume grows. Without an automated post-visit "
            f"SMS requesting a review, new reviews accumulate at whatever rate clients "
            f"volunteer them, which slows during busy periods. A structured automated request "
            f"sent within 24 hours of every appointment captures client enthusiasm at its "
            f"peak and keeps {name} consistently visible in {city} local search results."
        )
    else:
        return (
            f"With {reviews} reviews and a {rating:.1f}-star rating, {name} has a real "
            f"opportunity to improve its reputation score through more consistent review "
            f"collection. Without an automated post-visit review request, new reviews arrive "
            f"sporadically and the rating can stagnate even when the client experience is "
            f"strong. A systematic SMS request after every appointment would steadily improve "
            f"both volume and average score over the next 60 to 90 days."
        )


def gen_s05(name, city, ig_handle, rating):
    """05 — variable title and content based on signals"""
    if not ig_handle:
        title = "No Visible Social Media Presence"
        body  = (
            f"Without an active Instagram presence, {name} is invisible to a growing segment "
            f"of the {city} market that discovers aesthetics providers through social media "
            f"before ever searching on Google. Aesthetics clients often form their first "
            f"impression of a practice through its Instagram content, and the absence of a "
            f"profile means the practice is not participating in that discovery channel at all."
        )
    elif rating < 4.5:
        title = "Reputation Risk Without Active Management"
        body  = (
            f"A rating below 4.5 stars makes it harder for new clients in {city} to "
            f"confidently choose {name} over competitors with stronger reputations. Without a "
            f"systematic review management approach, the rating can remain stuck or decline "
            f"further as critical reviews carry more weight without consistent positive volume "
            f"to offset them. Active reputation management is the most direct path to "
            f"improving this metric and protecting the practice's ability to attract new clients."
        )
    else:
        title = "No Visible Retention System"
        body  = (
            f"There is no membership, prepaid treatment package, or automated loyalty "
            f"touchpoint visible in the current client experience at {name}. Aesthetics "
            f"clients who commit to a package or membership visit more frequently, spend more "
            f"per year, and refer more friends than those who book one appointment at a time. "
            f"Introducing even a basic monthly membership for recurring treatments would "
            f"meaningfully increase revenue predictability and client lifetime value."
        )
    return title, body


def gen_wins(name, city, reviews, ig_handle):
    """06 — Quick Wins (3 items)"""
    wins = []

    # Win 1: Reactivation
    if reviews > 200:
        wins.append((
            "Reactivate 90-day inactive clients",
            f"Send a personalized SMS to every client who has not visited in 90 days with a "
            f"message tied to their treatment cycle, such as a Botox refresh reminder or a "
            f"seasonal skin treatment suggestion. For a practice with {reviews} reviews "
            f"representing an established client base, these messages convert at an "
            f"exceptionally high rate because clients already have a strong positive "
            f"association with {name}.",
        ))
    else:
        wins.append((
            "Launch a client reactivation campaign",
            f"Send a warm personalized SMS to past clients who have not returned in 60 to 90 "
            f"days. Even a small list of lapsed clients represents significant recoverable "
            f"revenue, and a single well-timed message can bring several back on the books "
            f"within the week.",
        ))

    # Win 2: Review automation
    wins.append((
        "Automate post-visit review requests",
        f"Set up an automatic review request SMS within 24 hours of every appointment to "
        f"keep new reviews flowing consistently. Consistent review velocity improves local "
        f"search visibility and keeps {name} competitive in {city} area searches.",
    ))

    # Win 3: Lead capture or Instagram
    if not ig_handle:
        wins.append((
            "Establish an Instagram presence",
            f"Create an active Instagram profile and post consistently to reach {city} clients "
            f"who discover aesthetic providers through social media. Even a basic profile with "
            f"before/after results and treatment highlights captures interest from a segment "
            f"of the market that never searches Google.",
        ))
    else:
        wins.append((
            "Add SMS capture for non-bookers",
            f"Collect phone numbers from every visitor who shows interest but does not "
            f"schedule, then enroll them in a short automated follow-up sequence. Most "
            f"aesthetics clients need two or three touchpoints before booking their first "
            f"appointment, and this single addition recovers a meaningful percentage of leads "
            f"that currently disappear.",
        ))

    return wins


# ── HTML renderer ─────────────────────────────────────────────────────────────

def render_html(name, city, ig_handle, reviews, rating):
    city_label   = city if city else "Florida"
    ig_display   = ig_handle if ig_handle else ""
    rating_str   = f"{rating:.1f}"

    # Data pills
    pills = [
        f'<div class="data-pill"><span>Google Reviews</span>{reviews}</div>',
        f'<div class="data-pill"><span>Rating</span>{rating_str} ★</div>',
    ]
    if ig_display:
        pills.append(f'<div class="data-pill"><span>Instagram</span>{ig_display}</div>')
    pills_html = "\n    ".join(pills)

    s01 = gen_s01(name, city_label, reviews, rating)
    s02 = gen_s02(name, city_label, reviews)
    s03 = gen_s03(name, city_label, reviews, rating)
    s04 = gen_s04(name, city_label, reviews, rating)
    s05_title, s05 = gen_s05(name, city_label, ig_handle, rating)
    wins = gen_wins(name, city_label, reviews, ig_handle)

    wins_html = ""
    for i, (wt, wb) in enumerate(wins, 1):
        wins_html += (
            f'      <div class="win-card"><div class="win-number">{i}</div>'
            f'<div class="win-text"><strong>{wt}</strong>{wb}</div></div>\n'
        )

    cta_body = (
        f"If you'd like, I can show you the exact system other {city_label} "
        f"med spas use to bring back inactive clients."
    )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{name} \u2014 Digital Presence Audit | Auralith Digital</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=DM+Sans:opsz,wght@9..40,300;9..40,400;9..40,500;9..40,600&display=swap" rel="stylesheet">
<style>
  :root {{
    --ink: #0f0f0f; --ink-mid: #374151; --ink-light: #6b7280; --ink-faint: #e5e7eb;
    --surface: #ffffff; --surface-2: #f9fafb; --gold: #b5893a;
    --green: #16a34a; --green-light: #f0fdf4; --radius: 12px;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'DM Sans', sans-serif; background: var(--surface); color: var(--ink); line-height: 1.65; -webkit-font-smoothing: antialiased; }}
  .header {{ border-bottom: 1px solid var(--ink-faint); padding: 18px 40px; display: flex; align-items: center; justify-content: space-between; position: sticky; top: 0; background: rgba(255,255,255,0.96); backdrop-filter: blur(8px); z-index: 100; }}
  .logo {{ font-family: 'DM Serif Display', serif; font-size: 18px; color: var(--ink); }}
  .logo span {{ color: var(--gold); }}
  .header-meta {{ font-size: 11px; font-weight: 500; color: var(--ink-light); letter-spacing: 0.08em; text-transform: uppercase; }}
  .hero {{ padding: 60px 40px 40px; max-width: 820px; margin: 0 auto; opacity: 0; transform: translateY(14px); animation: fadeUp 0.5s ease 0.1s forwards; }}
  @keyframes fadeUp {{ to {{ opacity: 1; transform: none; }} }}
  .confidential {{ font-size: 11px; font-weight: 600; letter-spacing: 0.12em; text-transform: uppercase; color: var(--gold); margin-bottom: 10px; }}
  .business-name {{ font-family: 'DM Serif Display', serif; font-size: clamp(28px, 5vw, 44px); line-height: 1.15; margin-bottom: 10px; }}
  .hero-sub {{ font-size: 15px; color: var(--ink-light); margin-bottom: 32px; }}
  .data-row {{ display: flex; gap: 12px; flex-wrap: wrap; margin-bottom: 48px; }}
  .data-pill {{ background: var(--surface-2); border: 1px solid var(--ink-faint); border-radius: 100px; padding: 6px 16px; font-size: 13px; font-weight: 500; }}
  .data-pill span {{ color: var(--ink-light); font-weight: 400; margin-right: 4px; }}
  .main {{ max-width: 820px; margin: 0 auto; padding: 0 40px 80px; }}
  .section {{ margin-bottom: 40px; opacity: 0; transform: translateY(12px); animation: fadeUp 0.45s ease forwards; }}
  .section:nth-child(1) {{ animation-delay: 0.2s; }} .section:nth-child(2) {{ animation-delay: 0.3s; }}
  .section:nth-child(3) {{ animation-delay: 0.4s; }} .section:nth-child(4) {{ animation-delay: 0.5s; }}
  .section:nth-child(5) {{ animation-delay: 0.6s; }} .section:nth-child(6) {{ animation-delay: 0.7s; }}
  .section-label {{ font-size: 11px; font-weight: 600; letter-spacing: 0.1em; text-transform: uppercase; color: var(--ink-light); margin-bottom: 8px; display: flex; align-items: center; gap: 8px; }}
  .section-label::after {{ content: ''; flex: 1; height: 1px; background: var(--ink-faint); }}
  .card {{ background: var(--surface-2); border: 1px solid var(--ink-faint); border-radius: var(--radius); padding: 20px 24px; font-size: 14px; color: var(--ink-mid); }}
  .card p {{ margin-bottom: 8px; }} .card p:last-child {{ margin-bottom: 0; }} .card strong {{ color: var(--ink); }}
  .card-positive {{ background: var(--green-light); border-color: #bbf7d0; }}
  .card-positive::before {{ content: '\u2713'; display: inline-block; background: var(--green); color: white; width: 20px; height: 20px; border-radius: 50%; text-align: center; line-height: 20px; font-size: 11px; font-weight: 700; margin-bottom: 10px; }}
  .wins-grid {{ display: flex; flex-direction: column; gap: 12px; }}
  .win-card {{ display: flex; gap: 14px; align-items: flex-start; background: var(--surface-2); border: 1px solid var(--ink-faint); border-radius: var(--radius); padding: 16px 20px; }}
  .win-number {{ background: var(--ink); color: white; font-size: 13px; font-weight: 700; min-width: 28px; height: 28px; border-radius: 50%; display: flex; align-items: center; justify-content: center; flex-shrink: 0; margin-top: 2px; }}
  .win-text {{ font-size: 14px; color: var(--ink-mid); }}
  .win-text strong {{ display: block; color: var(--ink); font-size: 15px; margin-bottom: 3px; }}
  .cta-strip {{ background: var(--ink); color: white; text-align: center; padding: 56px 40px; }}
  .cta-title {{ font-family: 'DM Serif Display', serif; font-size: 28px; margin-bottom: 10px; }}
  .cta-sub {{ font-size: 14px; color: rgba(255,255,255,0.6); max-width: 480px; margin: 0 auto 24px; }}
  .cta-body {{ font-size: 15px; color: rgba(255,255,255,0.85); max-width: 460px; margin: 0 auto 24px; line-height: 1.7; }}
  .cta-btn {{ display: inline-block; background: var(--gold); color: white; font-size: 14px; font-weight: 600; padding: 13px 28px; border-radius: 8px; text-decoration: none; letter-spacing: 0.02em; }}
  .cta-ps {{ margin-top: 28px; font-size: 13px; color: rgba(255,255,255,0.45); max-width: 460px; margin-left: auto; margin-right: auto; line-height: 1.7; }}
  footer {{ border-top: 1px solid var(--ink-faint); padding: 24px 40px; display: flex; justify-content: space-between; align-items: center; font-size: 12px; color: var(--ink-light); }}
  .footer-logo {{ font-family: 'DM Serif Display', serif; font-size: 15px; color: var(--ink); }}
  .footer-logo span {{ color: var(--gold); }}
  @media (max-width: 600px) {{
    .header {{ padding: 14px 20px; }} .hero, .main {{ padding-left: 20px; padding-right: 20px; }}
    .win-card {{ flex-direction: column; gap: 8px; }} footer {{ flex-direction: column; gap: 8px; text-align: center; }}
  }}
</style>
</head>
<body>
<header class="header">
  <div class="logo">Auralith<span>.</span></div>
  <div class="header-meta">Digital Presence Audit &middot; {AUDIT_DATE}</div>
</header>
<div class="hero">
  <div class="confidential">Confidential Report &middot; {city_label}, FL</div>
  <h1 class="business-name">{name}</h1>
  <p class="hero-sub">We analyzed your online presence across Google, Instagram, and your booking experience to identify where patient revenue may be leaking.</p>
  <div class="data-row">
    {pills_html}
  </div>
</div>
<main class="main">
  <div class="section">
    <div class="section-label">01 \u2014 What\u2019s Working</div>
    <div class="card card-positive"><p>{s01}</p></div>
  </div>
  <div class="section">
    <div class="section-label">02 \u2014 Missed Bookings Without Follow-Up</div>
    <div class="card"><p>{s02}</p></div>
  </div>
  <div class="section">
    <div class="section-label">03 \u2014 Inactive Client Leakage</div>
    <div class="card"><p>{s03}</p></div>
  </div>
  <div class="section">
    <div class="section-label">04 \u2014 No Automated Review Generation</div>
    <div class="card"><p>{s04}</p></div>
  </div>
  <div class="section">
    <div class="section-label">05 \u2014 {s05_title}</div>
    <div class="card"><p>{s05}</p></div>
  </div>
  <div class="section">
    <div class="section-label">06 \u2014 Quick Wins</div>
    <div class="wins-grid">
{wins_html}    </div>
  </div>
</main>
<div class="cta-strip">
  <div class="cta-title">Curious how other med spas fix this?</div>
  <div class="cta-sub">Most clinics solve these issues with a few simple automations: automatic client reactivation, post-visit review requests, and SMS follow-ups for missed bookings.</div>
  <p class="cta-body">{cta_body}</p>
  <a href="mailto:hello@auralithdigital.com?subject=Show me the system" class="cta-btn">Just reply \u201cshow me\u201d</a>
  <p class="cta-ps"><strong style="color:rgba(255,255,255,0.6)">PS</strong> \u2014 I recorded this report manually for your clinic. If you\u2019d like, I can also show you the exact system med spas use to bring back inactive clients automatically.</p>
</div>
<footer>
  <div class="footer-logo">Auralith<span>.</span> Digital</div>
  <div>This report was prepared exclusively for {name} &middot; {AUDIT_DATE}</div>
</footer>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    # 1. Load already-done names from CSV
    csv_names = set()
    if os.path.exists(CSV_PATH):
        with open(CSV_PATH, newline='', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                n = row.get('name', '').strip()
                if n:
                    csv_names.add(n.lower())
    print(f"CSV businesses to skip: {len(csv_names)}")

    # 2. Load existing subdirectory slugs (already generated in new format)
    existing_slugs = set()
    if os.path.exists(OUTPUT_BASE):
        for entry in os.scandir(OUTPUT_BASE):
            if entry.is_dir():
                idx = os.path.join(entry.path, 'index.html')
                if os.path.exists(idx):
                    existing_slugs.add(entry.name)
    print(f"Existing subdirectory slugs: {len(existing_slugs)}")

    # 3. Read XLSX
    wb = openpyxl.load_workbook(XLSX_PATH)
    lt = wb['Lead Tracker']

    rows = []
    for row in range(4, 800):
        name = lt.cell(row=row, column=1).value
        if not name:
            break
        name    = str(name).strip()
        city    = str(lt.cell(row=row, column=2).value or '').strip()
        ig      = clean_ig(lt.cell(row=row, column=5).value)
        reviews = safe_int(lt.cell(row=row, column=12).value)
        rating  = safe_float(lt.cell(row=row, column=13).value)
        rows.append((reviews, name, city, ig, rating))

    print(f"Total XLSX rows: {len(rows)}")

    # 4. Sort descending by review count
    rows.sort(key=lambda x: x[0], reverse=True)

    # 5. Filter: skip CSV names and existing slugs
    candidates = []
    for reviews, name, city, ig, rating in rows:
        if name.lower() in csv_names:
            continue
        slug = slugify(name)
        if slug in existing_slugs:
            continue
        candidates.append((reviews, name, city, ig, rating, slug))

    print(f"Candidates after filtering: {len(candidates)}")

    # 6. Take first TARGET
    batch = candidates[:TARGET]
    print(f"Generating {len(batch)} reports...\n")

    generated = []
    for reviews, name, city, ig, rating, slug in batch:
        out_dir  = os.path.join(OUTPUT_BASE, slug)
        out_path = os.path.join(out_dir, 'index.html')
        os.makedirs(out_dir, exist_ok=True)
        html = render_html(name, city, ig, reviews, rating)
        with open(out_path, 'w', encoding='utf-8') as f:
            f.write(html)
        generated.append((slug, name, reviews, rating, city))
        print(f"  [{reviews:4d} rev | {rating:.1f}★] {name[:50]} -> {slug}/")

    print(f"\nDone. {len(generated)} reports written to {OUTPUT_BASE}/")
    return generated


if __name__ == '__main__':
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()
